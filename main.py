from tkinter import Tk, Canvas, Frame, Text, INSERT, Entry
from tkinter.constants import END

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Side, Border, NamedStyle, Alignment

from math import ceil

factor_of_font_size_to_width = {
    10: {
        "factor": 0.65,  # width / count of symbols at row
        "height": 15
    }
}


def get_height_for_row(sheet, row_number, font_size=12):
    font_params = factor_of_font_size_to_width[font_size]

    words_count_at_one_row = 54 * 0.5 / font_params["factor"]
    lines = ceil(len(str(sheet[f'X{row_number}'].value)) / words_count_at_one_row)
    height = lines * font_params["height"]

    return height


def get_info(sheet, i):
    date = sheet[f'A{i}'].value.replace('/', '.')
    partner = sheet[f'C{i}'].value \
        .replace('ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ', 'ООО') \
        .replace('Общество с ограниченной ответственностью', 'ООО') \
        .replace('ГРУППА КОМПАНИЙ', 'ГК') \
        .replace('Группа Компаний', 'ГК') \
        .replace('Индивидуальный предприниматель', 'ИП') \
        .replace('ИНДИВИДУАЛЬНЫЙ ПРЕДПРИНИМАТЕЛЬ', 'ИП') \
        .split('р/с')[0] \
        .split('Р/с')[0] \
        .split('Р/С')[0]
    if partner.split()[-1] == 'ООО':
        partner = 'ООО ' + partner[:-3]
    if partner.split()[-1] == '(ИП)':
        partner = 'ИП ' + partner[:-4]
    if partner.split()[-1] == 'ГКФХ)' and partner.split()[-2] == '(ИП':
        partner = 'ИП ГКФХ ' + partner[:-9]

    if partner.split()[0] == 'ИП' and partner.split()[1] == 'ГКФХ':
        partner = 'ИП ГКФХ' + partner[7:].title()
    elif partner.split()[0] == 'ИП':
        partner = 'ИП' + partner[2:].title()
    partner = partner.strip()
    money = sheet[f'D{i}'].value
    desc = sheet[f'E{i}'].value \
        .replace('/', '.') \
        .replace('\\', '.') \
        .replace(' января ', '.01.') \
        .replace(' февраля ', '.02.') \
        .replace(' марта ', '.03.') \
        .replace(' апреля ', '.04.') \
        .replace(' мая ', '.05.') \
        .replace(' июня ', '.06.') \
        .replace(' июля ', '.07.') \
        .replace(' августа ', '.08.') \
        .replace(' сентября ', '.09.') \
        .replace(' октября ', '.10.') \
        .replace(' ноября ', '.11.') \
        .replace(' декабря ', '.12.') \
        .replace('января', '.01.') \
        .replace('февраля', '.02.') \
        .replace('марта', '.03.') \
        .replace('апреля', '.04.') \
        .replace('мая', '.05.') \
        .replace('июня', '.06.') \
        .replace('июля', '.07.') \
        .replace('августа', '.08.') \
        .replace('сентября', '.09.') \
        .replace('октября', '.10.') \
        .replace('ноября', '.11.') \
        .replace('декабря', '.12.') \
        .replace(' Января ', '.01.') \
        .replace(' Февраля ', '.02.') \
        .replace(' Марта ', '.03.') \
        .replace(' Апреля ', '.04.') \
        .replace(' Мая ', '.05.') \
        .replace(' Июня ', '.06.') \
        .replace(' Июля ', '.07.') \
        .replace(' Августа ', '.08.') \
        .replace(' Сентября ', '.09.') \
        .replace(' Октября ', '.10.') \
        .replace(' Ноября ', '.11.') \
        .replace(' Декабря ', '.12.') \
        .replace('Января', '.01.') \
        .replace('Февраля', '.02.') \
        .replace('Марта', '.03.') \
        .replace('Апреля', '.04.') \
        .replace('Мая', '.05.') \
        .replace('Июня', '.06.') \
        .replace('Июля', '.07.') \
        .replace('Августа', '.08.') \
        .replace('Сентября', '.09.') \
        .replace('Октября', '.10.') \
        .replace('Ноября', '.11.') \
        .replace('Декабря', '.12.') \
        .replace(' ЯНВАРЯ ', '.01.') \
        .replace(' ФЕВРАЛЯ ', '.02.') \
        .replace(' МАРТА ', '.03.') \
        .replace(' АПРЕЛЯ ', '.04.') \
        .replace(' МАЯ ', '.05.') \
        .replace(' ИЮНЯ ', '.06.') \
        .replace(' ИЮЛЯ ', '.07.') \
        .replace(' АВГУСТА ', '.08.') \
        .replace(' СЕНТЯБРЯ ', '.09.') \
        .replace(' ОКТЯБРЯ ', '.10.') \
        .replace(' НОЯБРЯ ', '.11.') \
        .replace(' ДЕКАБРЯ ', '.12.') \
        .replace('ЯНВАРЯ', '.01.') \
        .replace('ФЕВРАЛЯ', '.02.') \
        .replace('МАРТА', '.03.') \
        .replace('АПРЕЛЯ', '.04.') \
        .replace('МАЯ', '.05.') \
        .replace('ИЮНЯ', '.06.') \
        .replace('ИЮЛЯ', '.07.') \
        .replace('АВГУСТА', '.08.') \
        .replace('СЕНТЯБРЯ', '.09.') \
        .replace('ОКТЯБРЯ', '.10.') \
        .replace('НОЯБРЯ', '.11.') \
        .replace('ДЕКАБРЯ', '.12.') \
        .replace(' январь ', '.01.') \
        .replace(' февраль ', '.02.') \
        .replace(' март ', '.03.') \
        .replace(' апрель ', '.04.') \
        .replace(' май ', '.05.') \
        .replace(' июнь ', '.06.') \
        .replace(' июль ', '.07.') \
        .replace(' август ', '.08.') \
        .replace(' сентябрь ', '.09.') \
        .replace(' октябрь ', '.10.') \
        .replace(' ноябрь ', '.11.') \
        .replace(' декабрь ', '.12.') \
        .replace('январь', '.01.') \
        .replace('февраль', '.02.') \
        .replace('март', '.03.') \
        .replace('апрель', '.04.') \
        .replace('май', '.05.') \
        .replace('июнь', '.06.') \
        .replace('июль', '.07.') \
        .replace('август', '.08.') \
        .replace('сентябрь', '.09.') \
        .replace('октябрь', '.10.') \
        .replace('ноябрь', '.11.') \
        .replace('декабрь', '.12.') \
        .replace(' Январь ', '.01.') \
        .replace(' Февраль ', '.02.') \
        .replace(' Март ', '.03.') \
        .replace(' Апрель ', '.04.') \
        .replace(' Май ', '.05.') \
        .replace(' Июнь ', '.06.') \
        .replace(' Июль ', '.07.') \
        .replace(' Август ', '.08.') \
        .replace(' Сентябрь ', '.09.') \
        .replace(' Октябрь ', '.10.') \
        .replace(' Ноябрь ', '.11.') \
        .replace(' Декабрь ', '.12.') \
        .replace('Январь', '.01.') \
        .replace('Февраль', '.02.') \
        .replace('Март', '.03.') \
        .replace('Апрель', '.04.') \
        .replace('Май', '.05.') \
        .replace('Июнь', '.06.') \
        .replace('Июль', '.07.') \
        .replace('Август', '.08.') \
        .replace('Сентябрь', '.09.') \
        .replace('Октябрь', '.10.') \
        .replace('Ноябрь', '.11.') \
        .replace('Декабрь', '.12.') \
        .replace(' ЯНВАРЬ ', '.01.') \
        .replace(' ФЕВРАЛЬ ', '.02.') \
        .replace(' МАРТ ', '.03.') \
        .replace(' АПРЕЛЬ ', '.04.') \
        .replace(' МАЙ ', '.05.') \
        .replace(' ИЮНЬ ', '.06.') \
        .replace(' ИЮЛЬ ', '.07.') \
        .replace(' АВГУСТ ', '.08.') \
        .replace(' СЕНТЯБРЬ ', '.09.') \
        .replace(' ОКТЯБРЬ ', '.10.') \
        .replace(' НОЯБРЬ ', '.11.') \
        .replace(' ДЕКАБРЬ ', '.12.') \
        .replace('ЯНВАРЬ', '.01.') \
        .replace('ФЕВРАЛЬ', '.02.') \
        .replace('МАРТ', '.03.') \
        .replace('АПРЕЛЬ', '.04.') \
        .replace('МАЙ', '.05.') \
        .replace('ИЮНЬ', '.06.') \
        .replace('ИЮЛЬ', '.07.') \
        .replace('АВГУСТ', '.08.') \
        .replace('СЕНТЯБРЬ', '.09.') \
        .replace('ОКТЯБРЬ', '.10.') \
        .replace('НОЯБРЬ', '.11.') \
        .replace('ДЕКАБРЬ', '.12.') \
        .replace(',', ' , ') \
        .replace('ЗА', '') \
        .replace('За', '') \
        .replace('за', '') \
        .replace('НА', '') \
        .replace('На', '') \
        .replace('на', '') \
        .replace('услуги', '') \
        .replace('Услуги', '') \
        .replace('УСЛУГИ', '') \
        .replace('ТРАНСПОРТНЫЕ', '') \
        .replace('Транспортные', '') \
        .replace('транспортные', '') \
        .replace('ОПЛАТА', '') \
        .replace('Оплата', '') \
        .replace('оплата', '') \
        .replace('ОПЛАТУ', '') \
        .replace('Оплату', '') \
        .replace('оплату', '') \
        .replace('№', ' № ') \
        .replace('N', ' N ') \
        .replace('#', ' # ') \
        .replace('г.', ' ') \
        .replace('Г.', ' ') \
        .replace('Г', ' ') \
        .replace('г', ' ') \
        .replace('.22', '.2022') \
        .split()
    bill = []
    j = 0
    while j < len(desc):
        try:
            if desc[j].lower()[:2] == 'сч':
                if desc[j + 1] == '№' or desc[j + 1] == 'N' or desc[j + 1] == '#':
                    bill.append([desc[j + 2], desc[j + 4]])
                    j += 4
                elif desc[j + 1][0] >= '0' and desc[j + 1][0] <= '9' or desc[j + 1][0].lower() == 'а' or desc[j + 1][0].lower() == 'a':
                    bill.append([desc[j + 1], desc[j + 3]])
                    j += 3
                j += 1
                while desc[j].lower()[0] == 'и' or desc[j].lower()[0] == ',':
                    if desc[j + 1].lower()[:2] == 'сч':
                        j += 1
                    if desc[j + 1] == '№' or desc[j + 1] == 'N' or desc[j + 1] == '#':
                        bill.append([desc[j + 2], desc[j + 4]])
                        j += 4
                    elif desc[j + 1][0] >= '0' and desc[j + 1][0] <= '9' or desc[j + 1][0].lower() == 'а' or desc[j + 1][0].lower() == 'a':
                        bill.append([desc[j + 1], desc[j + 3]])
                        j += 3
                    j += 1
                j -= 1
        except Exception:
            pass
        j += 1

    print(i, date, partner, money, bill)
    return [date, partner, money, bill, sheet[f'E{i}'].value]


def put_info(sheet, info, i):

    sheet.merge_cells(f'A{i + 6}:G{i + 6}')
    try:
        sheet[f'A{i + 6}'].style = text_cell_style
    except ValueError:
        pass
    sheet[f'A{i + 6}'].value = i

    sheet.merge_cells(f'H{i + 6}:W{i + 6}')
    try:
        sheet[f'H{i + 6}'].style = text_cell_style
    except ValueError:
        pass
    sheet[f'H{i + 6}'].value = info[0]

    sheet.merge_cells(f'X{i + 6}:BY{i + 6}')
    try:
        sheet[f'X{i + 6}'].style = text_cell_style
    except ValueError:
        pass
    bills = ''
    if len(info[3]) == 1:
        bills = f' по счету №{info[3][0][0]} от {info[3][0][1]}'
    elif len(info[3]) > 1:
        bills = ' по счетам'
        for bill in info[3]:
            bills += f' №{bill[0]},'
        bills = bills[:-1] + ' от'
        for bill in info[3]:
            bills += f' {bill[1]},'
        bills = bills[:-1]
    sheet[f'X{i + 6}'].value = f'услуги по перевозкам от {info[1]}' + bills

    sheet.row_dimensions[i + 6].height = get_height_for_row(sheet, i + 6, 10)
    # if len(sheet[f'X{i + 6}'].value) > 82:
    #     sheet.row_dimensions[i + 6].height = 45
    # else:
    #     sheet.row_dimensions[i + 6].height = 30

    sheet.merge_cells(f'BZ{i + 6}:CV{i + 6}')
    try:
        sheet[f'BZ{i + 6}'].style = money_cell_style
    except ValueError:
        pass
    sheet[f'BZ{i + 6}'].value = info[2]


def put_total(sheet, i):
    sheet.row_dimensions[i + 6].height = 15
    sheet.merge_cells(f'A{i + 6}:BY{i + 6}')
    try:
        sheet[f'A{i + 6}'].style = text_cell_style
    except ValueError:
        pass
    sheet[f'A{i + 6}'].value = 'Итого за налоговый период'
    sheet.merge_cells(f'BZ{i + 6}:CV{i + 6}')
    try:
        sheet[f'BZ{i + 6}'].style = money_cell_style
    except ValueError:
        pass
    sheet[f'BZ{i + 6}'] = f'=SUM(BZ7:CV{i + 5})'


def count(sheet):
    return len([row for row in range(1, sheet.max_row + 1)
                if sheet[f'A{row}'].value])


def add_style(book):
    text_cell_style = NamedStyle(name="TextCellStyle")
    thin = Side(style='thin')
    text_cell_style.border = Border(top=thin, left=thin, bottom=thin, right=thin)
    text_cell_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    try:
        book.add_named_style(text_cell_style)
    except ValueError:
        print('Text style is already applied')

    money_cell_style = NamedStyle(name="MoneyCellStyle")
    thin = Side(style='thin')
    money_cell_style.border = Border(top=thin, left=thin, bottom=thin, right=thin)
    money_cell_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    money_cell_style.number_format = '### ### ###"-00"'
    try:
        book.add_named_style(money_cell_style)
    except ValueError:
        print('Money style is already applied')

    return text_cell_style, money_cell_style


def get_user_info(event):
    new_bills = []
    temp = bills_text.get("1.0", END).split()
    if len(temp) % 2:
        print('Угораешь? Напиши нормально!!!!!')
    else:
        if len(temp):
            for k in range(1, len(temp) // 2 + 1):
                new_bills.append([temp[(k - 1) * 2], temp[(k - 1) * 2 + 1]])
            info[3] = new_bills
        root.destroy()


ans_book = load_workbook('template.xlsx')
ans_sheet = ans_book.worksheets[1]

stat_book = load_workbook('Events.xlsx')
stat_sheet = stat_book.worksheets[0]

text_cell_style, money_cell_style = add_style(ans_book)

count = count(stat_sheet)

for i in range(count, 1, -1):
    info = get_info(stat_sheet, i)
    root = Tk()
    root.title('Выберите верные данные')
    root.geometry('700x500')
    root.resizable(width=False, height=False)

    canvas = Canvas(root, height=700, width=500)
    canvas.pack()

    main_frame = Frame(root, bg='gray')
    main_frame.place(rely=0, relx=0, relheight=1, relwidth=1)

    desc_frame = Frame(main_frame)
    desc_frame.place(rely=0.03, relx=0.05, relheight=0.44, relwidth=0.9)

    bills_frame = Frame(main_frame)
    bills_frame.place(rely=0.53, relx=0.05, relheight=0.44, relwidth=0.9)

    # print(info[4])

    main_text = Text(desc_frame, bg='white')
    main_text.insert(INSERT, info[4])
    main_text.pack()

    bills_text = Text(bills_frame, bg='white')
    bills = ''
    for j, bill in enumerate(info[3]):
        # bills += f'{j + 1})' + bill[0] + ' от ' + bill[1] + '\n'
        bills += bill[0] + ' ' + bill[1] + '\n'
    bills_text.insert(INSERT, bills)
    bills_text.pack()

    root.bind('<Escape>', get_user_info)
    root.eval('tk::PlaceWindow . center')

    root.mainloop()

    print(info[3])

    put_info(ans_sheet, info, count + 1 - i)
put_total(ans_sheet, count)

ans_book.save('template.xlsx')
